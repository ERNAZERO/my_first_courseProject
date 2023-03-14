def info():
            show_info = int(input('''
    You are in information point
        1. Putty
        2. White paint
        3. Slate
    Select the material you want to know information about: '''))
            import openpyxl
            if show_info == 1:                             
                wb = openpyxl.load_workbook('stroishop.xlsx')
                wb.active = 0
                sheet = wb.active
                print ('\nThe price of putty(1pc) is: 200 KGS')
                print ('The amount of putty from the brand Glatt:', sheet['C2'].value,)
                print ('The amount of putty from the Milyi Dom:', sheet['C5'].value,)
                print ('Total:', sheet['C2'].value + sheet['C6'].value,'шт.')
                print ('The total cost of putty in stock is:',(200*(sheet['C2'].value + sheet['C5'].value)),'som' )
            
            elif show_info == 2:
                wb = openpyxl.load_workbook('stroishop.xlsx')
                wb.active = 0
                sheet = wb.active
                print ('\nThe price of white paint(1pc) is: 400 som')
                print ('The amount of white paint from the brand Glatt:', sheet['C3'].value)
                print ('The amount of white paint from the brand Milyi Dom:', sheet['C6'].value)
                print ('Tоtal:', sheet['C3'].value + sheet['C6'].value,'шт.')
                print ('The total cost of white paint in stock is:',(400*(sheet['C3'].value + sheet['C6'].value)),'sом' )

            elif show_info == 3:
                wb = openpyxl.load_workbook('stroishop.xlsx')
                wb.active = 0
                sheet = wb.active
                print ('\nThe price of the slate (1 piece) is: 300 som')
                print ('The amount of slate from the brand Glatt:', sheet['C4'].value)
                print ('The amount of slate from the brand Milyi Dom:', sheet['C7'].value)
                print ('Total:', sheet['C4'].value + sheet['C7'].value)
                print ('The total cost of slate in stock is:',(300*(sheet['C4'].value + sheet['C7'].value)),'soм' )