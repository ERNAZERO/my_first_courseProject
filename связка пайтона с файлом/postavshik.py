# from os import name, read, write
from numpy import product
from openpyxl import load_workbook
import pandas as pd


book = load_workbook('Поставщики.xlsx')
def add_pos(url, user_data):
    with pd.ExcelWriter(url, engine = 'openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        user_data.to_excel(writer, 'Поставщики', index=False)
        writer.save()



def read_post(url, sheet):
    read = pd.read_excel(url, sheet_name=sheet, index_col=False)
    return read



def add_postavshik(choice_in_menu):
    print ('\nFor add suppliers to the database, please enter his data by clicking the button   1\nIf you want to return to the menu click the button   0\n')
    
    ch = int(input(':'))

    if ch == 0:
        return choice_in_menu
    elif ch == 1:
        create_name_pos = input('\nName: ')
        create_surname_pos = input('Surname: ')
        create_address=input('Аddress: ')
        create_num=input('Number: ')
        tovar = input('Product: ')    
        saved_user = read_post('Поставщики.xlsx', 'Поставщики')
        local_user = pd.DataFrame(saved_user) 
        new_postavshik = local_user.append({'Name':create_name_pos, 'Surname':create_surname_pos, 'Address':create_address, 'Number':create_num, 'Product': tovar}, ignore_index=True)
        add_pos('Поставщики.xlsx', new_postavshik)
        print ('-----------------------------------------------------------------------------------------------------------------------------------')
        print ('\nSupplier added successfully!\n')
        print ('-----------------------------------------------------------------------------------------------------------------------------------')
    else:
        print ('There is no option {} here'.format(ch))
        return add_postavshik(choice_in_menu)






def delete_sup():
    from openpyxl import load_workbook
    just_show = pd.read_excel('Поставщики.xlsx', 'Поставщики')    
    print (just_show)

    wb = load_workbook('Поставщики.xlsx')
    ws = wb.active
    delete = int(input('\nChoice the index of the supplier you want to delete\n:'))
    delete += 2
    ws.delete_rows(delete)    
    print ('-----------------------------------------------------------------------------------------------------------------------------------')
    print ('The supplier has been removed')
    print ('-----------------------------------------------------------------------------------------------------------------------------------')
    wb.save ('Поставщики.xlsx')

# delete_sup()

























