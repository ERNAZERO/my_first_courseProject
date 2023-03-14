import pandas as pd
from colorama import Fore, Back
def dobav_material():
    from openpyxl import load_workbook

    wb = load_workbook('stroishop.xlsx')
    ws = wb.active
    naz = input ('\nName of material: ')
    brand = input ('Brand: ')
    cap = int(input('Capacity: '))
    ws.append([naz, brand, cap])
    print (Fore.GREEN + '\nProduct successfully added!' + Fore.WHITE)
    wb.save('stroishop.xlsx')


def Udal ():
    just_show = pd.read_excel('stroishop.xlsx', 'Количество строй материалов')    
    print (just_show)

    from openpyxl import load_workbook
    wb = load_workbook ('stroishop.xlsx')
    ws = wb.active
    delete_material = int(input(Fore.CYAN + '\nChoice the index of the product you want to delete\n:'))
    delete_material += 2
    ws.delete_rows(delete_material)    
    print (Fore.GREEN + '-----------------------------------------------------------------------------------------------------------------------------------\n')
    print (Fore. GREEN + 'The product has been removed')
    print (Fore.GREEN + '\n-----------------------------------------------------------------------------------------------------------------------------------')
    wb.save ('stroishop.xlsx')


